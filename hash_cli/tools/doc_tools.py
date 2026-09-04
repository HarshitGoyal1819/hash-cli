"""Document and data-file creation tools for the hash-cli agent.

Supports: Excel (.xlsx), PDF, YAML, CSV.
"""

from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any, Optional

from langchain_core.tools import tool


def _resolve(path: str) -> Path:
    p = Path(path).expanduser().resolve()
    p.parent.mkdir(parents=True, exist_ok=True)
    return p


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

@tool
def create_excel(
    path: str,
    sheets: dict,
    title: Optional[str] = None,
) -> str:
    """Create an Excel (.xlsx) workbook with one or more sheets.

    Args:
        path:   Output file path, e.g. "report.xlsx" or "output/data.xlsx".
        sheets: Dictionary mapping sheet name → list of rows.
                Each row is a list of cell values.
                The first row is treated as a header and will be bold.
                Example:
                  {
                    "Sales": [
                      ["Region", "Q1", "Q2"],
                      ["North",  1200, 1500],
                      ["South",   800,  950]
                    ]
                  }
        title:  Optional document title stored in workbook properties.

    Returns:
        Success message with the file path, or an error message.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        if title:
            wb.properties.title = title

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2D3A8C")  # dark blue
        header_align = Alignment(horizontal="center", vertical="center")

        for sheet_name, rows in sheets.items():
            ws = wb.create_sheet(title=sheet_name[:31])  # Excel 31-char limit

            for row_idx, row in enumerate(rows, start=1):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_align

            # Auto-fit column widths
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_width = 0
                for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        try:
                            max_width = max(max_width, len(str(cell.value or "")))
                        except Exception:
                            pass
                ws.column_dimensions[col_letter].width = min(max_width + 4, 60)

        target = _resolve(path)
        wb.save(str(target))
        sheet_names = ", ".join(sheets.keys())
        return f"✓ Excel file created: {target}  (sheets: {sheet_names})"

    except ImportError:
        return "Error: openpyxl not installed. Run: pip install openpyxl"
    except Exception as exc:
        return f"Error creating Excel file: {exc}"


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

@tool
def create_pdf(
    path: str,
    content: list,
    title: Optional[str] = None,
) -> str:
    """Create a PDF document from structured content.

    Args:
        path:    Output file path, e.g. "report.pdf".
        content: List of content blocks. Each block is a dict with a "type" key:

                 Title block:
                   {"type": "title", "text": "My Report"}

                 Heading block:
                   {"type": "heading", "text": "Section 1"}

                 Paragraph block:
                   {"type": "paragraph", "text": "Body text here..."}

                 Table block:
                   {"type": "table", "headers": ["Col A", "Col B"],
                    "rows": [["val1", "val2"], ["val3", "val4"]]}

                 Spacer block:
                   {"type": "spacer"}

        title:   Optional PDF metadata title.

    Returns:
        Success message with the file path, or an error message.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        )

        target = _resolve(path)
        doc = SimpleDocTemplate(
            str(target),
            pagesize=A4,
            rightMargin=2 * cm,
            leftMargin=2 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
            title=title or "",
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "HashTitle",
            parent=styles["Title"],
            fontSize=22,
            spaceAfter=18,
            textColor=colors.HexColor("#2D3A8C"),
        )
        heading_style = ParagraphStyle(
            "HashHeading",
            parent=styles["Heading2"],
            fontSize=14,
            spaceBefore=14,
            spaceAfter=8,
            textColor=colors.HexColor("#1E40AF"),
        )
        body_style = styles["BodyText"]
        body_style.fontSize = 11
        body_style.leading = 16

        story = []
        for block in content:
            btype = block.get("type", "paragraph")

            if btype == "title":
                story.append(Paragraph(block.get("text", ""), title_style))
                story.append(Spacer(1, 0.3 * cm))

            elif btype == "heading":
                story.append(Paragraph(block.get("text", ""), heading_style))

            elif btype == "paragraph":
                text = block.get("text", "").replace("\n", "<br/>")
                story.append(Paragraph(text, body_style))
                story.append(Spacer(1, 0.2 * cm))

            elif btype == "table":
                headers = block.get("headers", [])
                rows = block.get("rows", [])
                table_data = [headers] + rows if headers else rows

                tbl = Table(table_data, repeatRows=1)
                tbl.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2D3A8C")),
                    ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
                    ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE",   (0, 0), (-1, 0), 11),
                    ("ALIGN",      (0, 0), (-1, -1), "LEFT"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                     [colors.HexColor("#F3F4F6"), colors.white]),
                    ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
                    ("FONTSIZE",   (0, 1), (-1, -1), 10),
                    ("TOPPADDING",    (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("LEFTPADDING",   (0, 0), (-1, -1), 8),
                ]))
                story.append(tbl)
                story.append(Spacer(1, 0.4 * cm))

            elif btype == "spacer":
                story.append(Spacer(1, 0.5 * cm))

        doc.build(story)
        return f"✓ PDF created: {target}"

    except ImportError:
        return "Error: reportlab not installed. Run: pip install reportlab"
    except Exception as exc:
        return f"Error creating PDF: {exc}"


# ---------------------------------------------------------------------------
# YAML
# ---------------------------------------------------------------------------

@tool
def create_yaml(path: str, data: Any, comment: Optional[str] = None) -> str:
    """Write data to a YAML file.

    Args:
        path:    Output file path, e.g. "config.yaml" or "deploy/values.yaml".
        data:    Python dict or list to serialise as YAML.
        comment: Optional comment to place at the top of the file.

    Returns:
        Success message with the file path, or an error message.
    """
    try:
        import yaml

        target = _resolve(path)
        yaml_str = yaml.dump(data, default_flow_style=False, allow_unicode=True, sort_keys=False)

        if comment:
            header = "\n".join(f"# {line}" for line in comment.splitlines()) + "\n\n"
            yaml_str = header + yaml_str

        target.write_text(yaml_str, encoding="utf-8")
        return f"✓ YAML file created: {target}"

    except ImportError:
        return "Error: PyYAML not installed. Run: pip install PyYAML"
    except Exception as exc:
        return f"Error creating YAML file: {exc}"


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

@tool
def create_csv(
    path: str,
    headers: list,
    rows: list,
    delimiter: str = ",",
) -> str:
    """Write tabular data to a CSV file.

    Args:
        path:      Output file path, e.g. "data.csv".
        headers:   List of column header strings.
        rows:      List of rows; each row is a list of values.
        delimiter: Column delimiter. Default ",". Use "\\t" for TSV.

    Returns:
        Success message with row count and file path, or an error message.
    """
    try:
        target = _resolve(path)
        with target.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f, delimiter=delimiter)
            if headers:
                writer.writerow(headers)
            writer.writerows(rows)

        row_count = len(rows)
        return f"✓ CSV file created: {target}  ({row_count} data rows, {len(headers)} columns)"

    except Exception as exc:
        return f"Error creating CSV file: {exc}"
